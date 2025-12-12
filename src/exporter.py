import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from src.logger import setup_logger

logger = setup_logger()

def _flatten_data(data: dict) -> dict:
    """Helper to flatten the nested dictionary."""
    flat_data = {}
    
    # Seller
    seller = data.get('seller', {})
    flat_data['seller_name'] = seller.get('seller_name')
    flat_data['seller_address'] = seller.get('seller_address')
    flat_data['seller_phone'] = seller.get('seller_phone')
    
    # Buyer
    buyer = data.get('buyer', {})
    flat_data['buyer_name'] = buyer.get('buyer_name')
    flat_data['buyer_address'] = buyer.get('buyer_address')
    flat_data['buyer_phone'] = buyer.get('buyer_phone')
    flat_data['co_buyer_name'] = buyer.get('co_buyer_name')
    flat_data['co_buyer_address'] = buyer.get('co_buyer_address')
    flat_data['co_buyer_phone'] = buyer.get('co_buyer_phone')
    
    # Product
    product = data.get('product', {})
    flat_data['quantity'] = product.get('quantity')
    flat_data['items'] = product.get('items')
    flat_data['make_model'] = product.get('make_model')
    
    # Financial
    financial = data.get('financial', {})
    flat_data['apr_percentage'] = financial.get('apr_percentage')
    flat_data['finance_charge'] = financial.get('finance_charge')
    flat_data['amount_financed'] = financial.get('amount_financed')
    flat_data['total_of_payments'] = financial.get('total_of_payments')
    flat_data['total_sales_price'] = financial.get('total_sales_price')
    
    # Payment
    payment = data.get('payment', {})
    flat_data['number_of_payments'] = payment.get('number_of_payments')
    flat_data['payment_amount'] = payment.get('payment_amount')
    flat_data['payment_frequency'] = payment.get('payment_frequency')
    
    # Itemization
    itemization = data.get('itemization', {})
    flat_data['cash_price'] = itemization.get('cash_price')
    flat_data['down_payment'] = itemization.get('down_payment')
    flat_data['unpaid_balance'] = itemization.get('unpaid_balance')
    flat_data['amounts_paid_to_others'] = itemization.get('amounts_paid_to_others')
    flat_data['itemized_amount_financed'] = itemization.get('itemized_amount_financed')
    
    # Metadata
    metadata = data.get('metadata', {})
    flat_data['source_file'] = metadata.get('source_file')
    flat_data['extraction_date'] = metadata.get('extraction_date')
    
    return flat_data

def export_to_csv(data: dict, output_path: str) -> None:
    """Exports data to CSV."""
    try:
        flat_data = _flatten_data(data)
        fieldnames = list(flat_data.keys())
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerow(flat_data)
        logger.info(f"Successfully exported CSV to {output_path}")
    except Exception as e:
        logger.error(f"Failed to export CSV: {e}")
        raise

def export_to_excel(data: dict, output_path: str) -> None:
    """
    Exports data to formatted Excel file.
    """
    try:
        flat_data = _flatten_data(data)
        headers = list(flat_data.keys())
        values = list(flat_data.values())
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Write Header
        ws.append(headers)
        
        # Write Values
        ws.append(values)
        
        # Styling
        # Header Bold
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Currency Formatting and Column Widths
        currency_cols = [
            'finance_charge', 'amount_financed', 'total_of_payments', 'total_sales_price',
            'payment_amount', 'cash_price', 'down_payment', 'unpaid_balance',
            'amounts_paid_to_others', 'itemized_amount_financed'
        ]
        percentage_cols = ['apr_percentage']
        
        for idx, col_name in enumerate(headers, 1):
            # Auto-width (approximate)
            ws.column_dimensions[chr(64 + idx) if idx <= 26 else 'A'+chr(64+idx-26)].width = 20
            
            # Format Data Row (Row 2)
            cell = ws.cell(row=2, column=idx)
            if col_name in currency_cols and isinstance(cell.value, (int, float)):
                 cell.number_format = '$#,##0.00'
            elif col_name in percentage_cols and isinstance(cell.value, (int, float)):
                 cell.number_format = '0.00%'

        wb.save(output_path)
        logger.info(f"Successfully exported Excel to {output_path}")
    except Exception as e:
        logger.error(f"Failed to export Excel: {e}")
        raise
